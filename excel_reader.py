# excel_reader.py
from openpyxl import load_workbook
import os

def read_excel_sheets(folder_path):
    data = {}
    # Iterate over every Excel file in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            # Load the workbook and iterate over each sheet
            workbook = load_workbook(os.path.join(folder_path, filename))
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Assuming the first row is headers
                columns = [cell.value for cell in sheet[1]]
                rows = list(sheet.iter_rows(min_row=2, values_only=True))
                # Use a tuple (filename, sheet_name) as a key
                data[(filename, sheet_name)] = {'columns': columns, 'rows': rows}
    return data
